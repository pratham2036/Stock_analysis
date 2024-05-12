import os,dotenv,requests,datetime
from dotenv import find_dotenv,load_dotenv
from openpyxl import load_workbook,Workbook
dotenv_path=find_dotenv()
load_dotenv(dotenv_path)
API_KEY=os.getenv("API_KEY")

user_input = input("Enter Stock: ")
wb = load_workbook(filename='Book1.xlsx')
ws = wb.active
ws['A1'].value='Stock-ticker'
ws['B1'].value='Date'
ws['C1'].value='Eps'
ws['D1'].value='Revenue'
ws['E1'].value='EPS_YOY_growth'

ws['F1'].value='Revenue_YOY_growth'
ws['G1'].value='Market_capital'
ws['H1'].value='Short_intrest'
ws['A2'].value=''
ws['B2'].value=''
ws['C2'].value=''
ws['D2'].value=''
ws['E2'].value=''

ws['F2'].value=''
ws['G2'].value=''
ws['H2'].value=''

wb.save('Book1.xlsx')
    
def get_start_of_current_week():
    today = datetime.date.today()
    start_of_week = today - datetime.timedelta(days=today.weekday())
    return start_of_week

# Function to get the current date
def get_current_date():
    return datetime.date.today()

# Example usage
start_of_week = get_start_of_current_week()
current_date = get_current_date()
def get_earnings_data(symbol):
        
    url=f'https://financialmodelingprep.com/api/v3/earning_calendar?from={start_of_week}&to={current_date}&apikey={API_KEY}'
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for 4xx or 5xx errors
        data = response.json()
        for earning in data:
            if earning['symbol'] == symbol:
                curr_eps=earning['eps']
                curr_reve=earning['revenue']
                wb = load_workbook(filename='Book1.xlsx')
                ws = wb.active
                ws['A2'].value=earning['symbol']
                ws['B2'].value=earning['date']
                ws['C2'].value=earning['eps']
                ws['D2'].value=earning['revenue']
                
                wb.save('Book1.xlsx')
               
    except requests.exceptions.RequestException as e:
        print("Failed to fetch data for the current week:", e) 

    url=f'https://financialmodelingprep.com/api/v3/market-capitalization/{symbol}?apikey={API_KEY}'
    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for 4xx or 5xx errors
        data = response.json()
        for earning in data:
            if earning['symbol'] == symbol:
                wb = load_workbook(filename='Book1.xlsx')
                ws = wb.active
                ws['G2'].value=earning['marketCap']               
                wb.save('Book1.xlsx')
               
    except requests.exceptions.RequestException as e:
        print("Failed to fetch data for the current week:", e)    
                


# Function to fetch income statement data from Financial Modeling Prep API
def get_income_statement_data(symbol):
    url = f"https://financialmodelingprep.com/api/v3/income-statement/AAPL?apikey={API_KEY}&limit=10"
    response = requests.get(url)
    data = response.json()
    return data

# Function to calculate YOY revenue growth
def calculate_yoy_revenue_growth(income_statement_data):
    yoy_growth = []
    final_growth = 100  # Initial growth rate starts at 100%

    for i in range(len(income_statement_data)-1):
        current_revenue = income_statement_data[i]['revenue']
        previous_revenue = income_statement_data[i+1]['revenue']
        growth = ((current_revenue - previous_revenue) / previous_revenue) * 100
        final_growth *= (1 + growth / 100)  # Compound the growth
    final_growth -= 100 
    return final_growth

# Function to calculate YOY EPS growth
def calculate_yoy_eps_growth(income_statement_data):
    yoy_growth = []
    final_growth = 100
    for i in range(len(income_statement_data)-1):
        current_eps = income_statement_data[i]['eps']
        previous_eps = income_statement_data[i+1]['eps']
        growth = ((current_eps - previous_eps) / previous_eps) * 100
        final_growth *= (1 + growth / 100)  # Compound the growth
    final_growth -= 100 
    return final_growth


# Main function
def main():

    get_earnings_data(user_input)
    income_statement_data = get_income_statement_data(user_input)

    if income_statement_data:
        revenue_growth = calculate_yoy_revenue_growth(income_statement_data)
        eps_growth = calculate_yoy_eps_growth(income_statement_data)
        wb = load_workbook(filename='Book1.xlsx')
        ws = wb.active
        ws['F2'].value=revenue_growth
        ws['E2'].value=eps_growth
        wb.save('Book1.xlsx')
       
      

    else:
        print("Failed to retrieve income statement data.")
    wb = load_workbook(filename='Book1.xlsx')
    ws = wb.active
    if(ws['A1'].value=='symbol'):
        print("Succesfully uploaded in xlsx")
               
    wb.save('Book1.xlsx')

if __name__ == "__main__":
    main()
